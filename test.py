#!/usr/bin/python3

import re
import openpyxl

path = "100_multichoice.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

max_col = sheet_obj.max_column
m_row = sheet_obj.max_row

for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 8)
    print(cell_obj.value)
