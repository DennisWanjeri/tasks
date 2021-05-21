#!/usr/bin/python3

import re
import openpyxl

path = "100_multichoice.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

max_col = sheet_obj.max_column
m_row = sheet_obj.max_row
    
def cleanhtml(raw_html):
    cleanr = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});')
    clean_text = re.sub(cleanr, '', raw_html)
    return clean_text

for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 8)
    raw_data = cell_obj.value
    newstr = ""
    newstr = cleanhtml(newstr)
    cell_obj.value = newstr

wb_obj.save(filename="100_multichoice.xlsx")
